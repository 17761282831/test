import os

import openpyxl


class ExcelUtil:
    # 获得项目路径
    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    # 读取文件内容
    def read_excel(self):
        # 目前读取excel的第三方模块有openpyxl,xlrd，两种，一般使用openpyxl
        # 加载excel工作簿
        wb = openpyxl.load_workbook(self.get_object_path() + "data/login_data.xlsx")
        # 获得excel中的sheet对象,需要注意：sheet对象的方法不会自动补齐
        sheet = wb["login"]
        # 获取excel的值，使用sheet.cell().value方法
        data_list = []
        for rows in range(2, sheet.max_row + 1):
            temp_list = []
            for cols in range(1, sheet.max_column + 1):
                temp_list.append(sheet.cell(rows, cols).value)
            data_list.append(temp_list)
        return data_list
